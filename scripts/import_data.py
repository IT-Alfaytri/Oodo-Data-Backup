"""
Odoo Data Import Script (fast / Postgres)
=========================================
Reads the organized category workbooks and loads them into the Supabase tables
over a DIRECT Postgres connection: one instant TRUNCATE, then batched INSERTs via
psycopg2.execute_values. Reloads in minutes instead of the ~hour the old REST
(anon/service-key) importer took.

Config is via environment variables (never hardcode credentials):
    SUPABASE_DB_URL      postgresql://postgres:<pw>@db.<ref>.supabase.co:5432/postgres
                         (URL-encode special chars in the password, e.g. @ -> %40)
    ODOO_ORGANIZED_DIR   folder with the 01_*.xlsx … category workbooks
                         (default: D:\\ERP\\oodo codes\\odoo_organized)

Usage:
    pip install -r requirements.txt
    SUPABASE_DB_URL=... ODOO_ORGANIZED_DIR=... python import_data.py
"""

import os
import sys
import time
from datetime import datetime, date
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Fast path: load directly over Postgres (TRUNCATE + batched INSERT via
# execute_values) — minutes instead of the ~hour the REST importer took.
# Connection string comes from the environment only (never hardcoded / committed):
#   SUPABASE_DB_URL=postgresql://postgres:<pw>@db.<ref>.supabase.co:5432/postgres
DATABASE_URL = os.environ.get("SUPABASE_DB_URL", "")
if not DATABASE_URL:
    sys.exit("Set SUPABASE_DB_URL (postgresql://…) in the environment before running.")

# Organized category workbooks produced by the export pipeline (03_organize).
EXCEL_DIR = Path(os.environ.get("ODOO_ORGANIZED_DIR", r"D:\ERP\oodo codes\odoo_organized"))


def pg_insert(cur, table, cols, batch):
    """Fast batched INSERT via execute_values (one round-trip per page)."""
    if not batch:
        return
    sql = f'INSERT INTO {table} ({", ".join(cols)}) VALUES %s'
    execute_values(cur, sql, [tuple(r.get(c) for c in cols) for r in batch], page_size=1000)

BATCH_SIZE = 2000
PROGRESS_EVERY = 10000

# Foreign-key columns that need name->id resolution.
# Maps (db_column, child_table) -> (excel_file, sheet, name_column) of the parent.
# When the Excel cell is a string (display name), we look up the integer id from
# the parent sheet.  When it's already numeric, we use it directly.
FK_LOOKUPS = {
    ("order_id", "sale_order_lines"):       ("01_Sales.xlsx", "Sales Orders", "name"),
    ("order_id", "purchase_order_lines"):    ("02_Purchase.xlsx", "Purchase Orders", "name"),
    ("product_tmpl_id", "product_variants"): ("03_Products.xlsx", "Product Templates", "name"),
    ("move_id", "invoice_lines"):            ("06_Invoices.xlsx", "Invoices & Bills", "name"),
    ("full_reconcile_id", "partial_reconcile"): ("07_Reconciliation.xlsx", "Full Reconcile", "name"),
    ("picking_id", "stock_moves"):           ("09_Stock_Movements.xlsx", "Transfers", "name"),
    ("move_id", "stock_move_lines"):         ("09_Stock_Movements.xlsx", "Stock Moves", "name"),
    ("cost_id", "landed_cost_lines"):        ("10_Costing.xlsx", "Landed Costs", "name"),
    ("cost_id", "valuation_adjustments"):    ("10_Costing.xlsx", "Landed Costs", "name"),
}

# Set of all FK column names (for quick membership test)
FK_COLUMNS = {k[0] for k in FK_LOOKUPS}

# ---------------------------------------------------------------------------
# Import map: 30 sheet-to-table mappings across 13 Excel files
# ---------------------------------------------------------------------------

IMPORT_MAP = [
    # 01_Sales.xlsx
    {
        "excel_file": "01_Sales.xlsx",
        "sheet": "Sales Orders",
        "table": "sale_orders",

        "columns": {
            "id": "id",
            "name": "name",
            "state": "state",
            "partner_id": "partner_id",
            "date_order": "date_order",
            "amount_untaxed": "amount_untaxed",
            "amount_tax": "amount_tax",
            "amount_total": "amount_total",
            "invoice_status": "invoice_status",
            "user_id": "user_id",
            "team_id": "team_id",
            "warehouse_id": "warehouse_id",
            "margin": "margin",
            "margin_percent": "margin_percent",
            "discount_type": "discount_type",
            "discount_amount": "discount_amount",
            "currency_id": "currency_id",
            "company_id": "company_id",
            "x_studio_lpo_reference": "x_studio_lpo_reference_2",
            "x_studio_amount_in_words": "x_studio_amount_in_words",
        },
    },
    {
        "excel_file": "01_Sales.xlsx",
        "sheet": "Order Lines",
        "table": "sale_order_lines",

        "columns": {
            "id": "id",
            "order_id": "order_id",
            "order_name": "order_id",
            "product_id": "product_id",
            "product_uom_qty": "product_uom_qty",
            "price_unit": "price_unit",
            "discount": "discount",
            "price_subtotal": "price_subtotal",
            "price_total": "price_total",
            "qty_delivered": "qty_delivered",
            "qty_invoiced": "qty_invoiced",
            "margin": "margin",
            "margin_percent": "margin_percent",
            "purchase_price": "purchase_price",
            "state": "state",
        },
    },
    # 02_Purchase.xlsx
    {
        "excel_file": "02_Purchase.xlsx",
        "sheet": "Purchase Orders",
        "table": "purchase_orders",

        "columns": {
            "id": "id",
            "name": "name",
            "state": "state",
            "partner_id": "partner_id",
            "date_order": "date_order",
            "date_approve": "date_approve",
            "amount_untaxed": "amount_untaxed",
            "amount_tax": "amount_tax",
            "amount_total": "amount_total",
            "invoice_status": "invoice_status",
            "receipt_status": "receipt_status",
            "discount_type": "discount_type",
            "discount_amount": "discount_amount",
            "currency_id": "currency_id",
            "company_id": "company_id",
            "x_studio_amount_in_words": "x_studio_amount_in_words",
        },
    },
    {
        "excel_file": "02_Purchase.xlsx",
        "sheet": "Order Lines",
        "table": "purchase_order_lines",
        "columns": {
            "id": "id",
            "order_id": "order_id",
            "product_id": "product_id",
            "product_qty": "product_qty",
            "price_unit": "price_unit",
            "price_subtotal": "price_subtotal",
            "price_total": "price_total",
            "qty_received": "qty_received",
            "qty_invoiced": "qty_invoiced",
            "discount": "discount",
            "state": "state",
        },
    },
    # 03_Products.xlsx
    {
        "excel_file": "03_Products.xlsx",
        "sheet": "Product Templates",
        "table": "product_templates",
        "columns": {
            "id": "id",
            "name": "name",
            "type": "type",
            "categ_id": "categ_id",
            "list_price": "list_price",
            "standard_price": "standard_price",
            "uom_id": "uom_id",
            "qty_available": "qty_available",
            "virtual_available": "virtual_available",
            "tracking": "tracking",
            "cost_method": "cost_method",
            "valuation": "valuation",
            "sale_ok": "sale_ok",
            "purchase_ok": "purchase_ok",
            "active": "active",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "03_Products.xlsx",
        "sheet": "Product Variants",
        "table": "product_variants",
        "columns": {
            "id": "id",
            "product_tmpl_id": "product_tmpl_id",
            "name": "name",
            "barcode": "barcode",
            "default_code": "default_code",
            "qty_available": "qty_available",
            "standard_price": "standard_price",
        },
    },
    {
        "excel_file": "03_Products.xlsx",
        "sheet": "Categories",
        "table": "product_categories",
        "columns": {
            "id": "id",
            "name": "name",
            "complete_name": "complete_name",
            "parent_id": "parent_id",
            "product_count": "product_count",
        },
    },
    # 04_Contacts.xlsx
    {
        "excel_file": "04_Contacts.xlsx",
        "sheet": "Contacts",
        "table": "contacts",
        "columns": {
            "id": "id",
            "name": "name",
            "is_company": "is_company",
            "type": "type",
            "street": "street",
            "city": "city",
            "country_id": "country_id",
            "email": "email",
            "phone": "phone",
            "mobile": "mobile",
            "customer_rank": "customer_rank",
            "supplier_rank": "supplier_rank",
            "credit": "credit",
            "debit": "debit",
            "total_invoiced": "total_invoiced",
            "credit_limit": "credit_limit",
            "company_id": "company_id",
            "parent_id": "parent_id",
            "active": "active",
        },
    },
    # 05_Accounting.xlsx
    {
        "excel_file": "05_Accounting.xlsx",
        "sheet": "Chart of Accounts",
        "table": "chart_of_accounts",
        "columns": {
            "id": "id",
            "name": "name",
            "code": "code",
            "account_type": "account_type",
            "internal_group": "internal_group",
            "reconcile": "reconcile",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "05_Accounting.xlsx",
        "sheet": "Journals",
        "table": "journals",
        "columns": {
            "id": "id",
            "name": "name",
            "code": "code",
            "type": "type",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "05_Accounting.xlsx",
        "sheet": "Payments",
        "table": "payments",
        "columns": {
            "id": "id",
            "name": "name",
            "date": "date",
            "state": "state",
            "payment_type": "payment_type",
            "partner_type": "partner_type",
            "partner_id": "partner_id",
            "amount": "amount",
            "currency_id": "currency_id",
            "journal_id": "journal_id",
            "ref": "ref",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "05_Accounting.xlsx",
        "sheet": "Bank Statements",
        "table": "bank_statement_lines",
        "columns": {
            "id": "id",
            "move_id": "move_id",
            "partner_id": "partner_id",
            "payment_ref": "payment_ref",
            "amount": "amount",
            "is_reconciled": "is_reconciled",
        },
    },
    # 06_Invoices.xlsx
    {
        "excel_file": "06_Invoices.xlsx",
        "sheet": "Invoices & Bills",
        "table": "invoices",
        "columns": {
            "id": "id",
            "name": "name",
            "date": "date",
            "state": "state",
            "move_type": "move_type",
            "partner_id": "partner_id",
            "amount_untaxed": "amount_untaxed",
            "amount_tax": "amount_tax",
            "amount_total": "amount_total",
            "amount_residual": "amount_residual",
            "payment_state": "payment_state",
            "invoice_origin": "invoice_origin",
            "invoice_date": "invoice_date",
            "invoice_date_due": "invoice_date_due",
            "journal_id": "journal_id",
            "currency_id": "currency_id",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "06_Invoices.xlsx",
        "sheet": "Invoice Lines",
        "table": "invoice_lines",
        "columns": {
            "id": "id",
            "move_id": "move_id",
            "move_name": "move_name",
            "date": "date",
            "parent_state": "parent_state",
            "account_id": "account_id",
            "partner_id": "partner_id",
            "product_id": "product_id",
            "name": "name",
            "quantity": "quantity",
            "price_unit": "price_unit",
            "discount": "discount",
            "debit": "debit",
            "credit": "credit",
            "balance": "balance",
            "display_type": "display_type",
            "matching_number": "matching_number",
            "currency_id": "currency_id",
            "journal_id": "journal_id",
            "company_id": "company_id",
        },
    },
    # 07_Reconciliation.xlsx
    {
        "excel_file": "07_Reconciliation.xlsx",
        "sheet": "Full Reconcile",
        "table": "full_reconcile",
        "columns": {
            "id": "id",
            "name": "name",
        },
    },
    {
        "excel_file": "07_Reconciliation.xlsx",
        "sheet": "Partial Reconcile",
        "table": "partial_reconcile",
        "columns": {
            "id": "id",
            "debit_move_id": "debit_move_id",
            "credit_move_id": "credit_move_id",
            "full_reconcile_id": "full_reconcile_id",
            "amount": "amount",
            "max_date": "max_date",
        },
    },
    # 08_Inventory.xlsx
    {
        "excel_file": "08_Inventory.xlsx",
        "sheet": "Warehouses",
        "table": "warehouses",
        "columns": {
            "id": "id",
            "name": "name",
            "code": "code",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "08_Inventory.xlsx",
        "sheet": "Locations",
        "table": "stock_locations",
        "columns": {
            "id": "id",
            "name": "name",
            "complete_name": "complete_name",
            "usage": "usage",
            "warehouse_id": "warehouse_id",
        },
    },
    {
        "excel_file": "08_Inventory.xlsx",
        "sheet": "Stock On Hand",
        "table": "stock_quants",
        "columns": {
            "id": "id",
            "product_id": "product_id",
            "location_id": "location_id",
            "quantity": "quantity",
            "reserved_quantity": "reserved_quantity",
            "inventory_quantity": "inventory_quantity",
            "lot_id": "lot_id",
            "in_date": "in_date",
            "company_id": "company_id",
        },
    },
    # 09_Stock_Movements.xlsx
    {
        "excel_file": "09_Stock_Movements.xlsx",
        "sheet": "Transfers",
        "table": "stock_pickings",
        "columns": {
            "id": "id",
            "name": "name",
            "origin": "origin",
            "partner_id": "partner_id",
            "picking_type_id": "picking_type_id",
            "location_id": "location_id",
            "location_dest_id": "location_dest_id",
            "date": "date",
            "date_done": "date_done",
            "state": "state",
            "scheduled_date": "scheduled_date",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "09_Stock_Movements.xlsx",
        "sheet": "Stock Moves",
        "table": "stock_moves",
        "columns": {
            "id": "id",
            "name": "name",
            "date": "date",
            "product_id": "product_id",
            "product_uom": "product_uom",
            "product_uom_qty": "product_uom_qty",
            "quantity_done": "quantity_done",
            "location_id": "location_id",
            "location_dest_id": "location_dest_id",
            "picking_id": "picking_id",
            "origin": "origin",
            "state": "state",
            "price_unit": "price_unit",
            "partner_id": "partner_id",
            "warehouse_id": "warehouse_id",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "09_Stock_Movements.xlsx",
        "sheet": "Move Details",
        "table": "stock_move_lines",
        "columns": {
            "id": "id",
            "move_id": "move_id",
            "product_id": "product_id",
            "qty_done": "qty_done",
            "location_id": "location_id",
            "location_dest_id": "location_dest_id",
            "reference": "reference",
        },
    },
    {
        "excel_file": "09_Stock_Movements.xlsx",
        "sheet": "Scrap Orders",
        "table": "stock_scraps",
        "columns": {
            "id": "id",
            "name": "name",
            "product_id": "product_id",
            "scrap_qty": "scrap_qty",
            "scrap_location_id": "scrap_location_id",
            "date_done": "date_done",
        },
    },
    # 10_Costing.xlsx
    {
        "excel_file": "10_Costing.xlsx",
        "sheet": "Landed Costs",
        "table": "landed_costs",
        "columns": {
            "id": "id",
            "name": "name",
            "date": "date",
            "state": "state",
            "amount_total": "amount_total",
            "vendor_bill_id": "vendor_bill_id",
            "account_journal_id": "account_journal_id",
        },
    },
    {
        "excel_file": "10_Costing.xlsx",
        "sheet": "Landed Cost Lines",
        "table": "landed_cost_lines",
        "columns": {
            "id": "id",
            "cost_id": "cost_id",
            "name": "name",
            "product_id": "product_id",
            "price_unit": "price_unit",
            "split_method": "split_method",
            "account_id": "account_id",
        },
    },
    {
        "excel_file": "10_Costing.xlsx",
        "sheet": "Valuation Layers",
        "table": "valuation_layers",
        "columns": {
            "id": "id",
            "product_id": "product_id",
            "quantity": "quantity",
            "unit_cost": "unit_cost",
            "value": "value",
            "remaining_qty": "remaining_qty",
            "remaining_value": "remaining_value",
            "stock_move_id": "stock_move_id",
            "stock_landed_cost_id": "stock_landed_cost_id",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "10_Costing.xlsx",
        "sheet": "Valuation Adjustments",
        "table": "valuation_adjustments",
        "columns": {
            "id": "id",
            "name": "name",
            "cost_id": "cost_id",
            "product_id": "product_id",
            "former_cost": "former_cost",
            "additional_landed_cost": "additional_landed_cost",
            "final_cost": "final_cost",
        },
    },
    # 11_Manufacturing.xlsx
    {
        "excel_file": "11_Manufacturing.xlsx",
        "sheet": "BOM Lines",
        "table": "bom_lines",
        "columns": {
            "id": "id",
            "product_id": "product_id",
            "product_qty": "product_qty",
            "bom_id": "bom_id",
            "cost_share": "cost_share",
        },
    },
    # 12_HR.xlsx
    {
        "excel_file": "12_HR.xlsx",
        "sheet": "Employees",
        "table": "employees",
        "columns": {
            "id": "id",
            "name": "name",
            "department_id": "department_id",
            "job_title": "job_title",
            "company_id": "company_id",
        },
    },
    {
        "excel_file": "12_HR.xlsx",
        "sheet": "Departments",
        "table": "departments",
        "columns": {
            "id": "id",
            "name": "name",
            "company_id": "company_id",
        },
    },
    # 13_Approvals.xlsx
    {
        "excel_file": "13_Approvals.xlsx",
        "sheet": "Approval Requests",
        "table": "approval_requests",
        "columns": {
            "id": "id",
            "name": "name",
            "category_id": "category_id",
            "request_status": "request_status",
            "date": "date",
        },
    },
]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def build_fk_lookup(excel_file: str, sheet_name: str, name_col: str) -> dict:
    """Build a {name_string -> integer_id} map from a parent table's Excel sheet."""
    path = EXCEL_DIR / excel_file
    if not path.exists():
        print(f"  WARNING: FK lookup file not found: {path}")
        return {}

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: FK lookup sheet '{sheet_name}' not found in {excel_file}")
        wb.close()
        return {}

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows()
    header = next(rows_iter)
    headers = [cell.value for cell in header]

    id_idx = headers.index("id") if "id" in headers else None
    name_idx = headers.index(name_col) if name_col in headers else None
    if id_idx is None or name_idx is None:
        print(f"  WARNING: FK lookup missing 'id' or '{name_col}' in {excel_file}/{sheet_name}")
        wb.close()
        return {}

    lookup = {}
    for row in rows_iter:
        vals = [cell.value for cell in row]
        row_id = vals[id_idx] if id_idx < len(vals) else None
        row_name = vals[name_idx] if name_idx < len(vals) else None
        if row_id is not None and row_name is not None:
            try:
                lookup[str(row_name)] = int(row_id)
            except (ValueError, TypeError):
                pass
    wb.close()
    return lookup


def flatten_value(val):
    """Convert Odoo field values to plain Python types.

    - Many2one [id, 'name'] -> 'name'
    - False / None -> None
    - datetime/date -> ISO string
    - Everything else passes through
    """
    if val is False or val is None:
        return None
    if isinstance(val, list) and len(val) == 2:
        # Many2one: [id, 'display_name']
        return str(val[1])
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return val


def extract_fk_id(val, lookup: dict = None):
    """Extract integer ID from an FK value.

    Handles: numeric values, [id, 'name'] lists, and string display names
    (resolved via the optional lookup dict).
    """
    if val is False or val is None:
        return None
    if isinstance(val, list) and len(val) == 2:
        return int(val[0])
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(val)
    except (ValueError, TypeError):
        pass
    if lookup and isinstance(val, str):
        return lookup.get(val)
    return None


def import_sheet(config: dict, cur) -> int:
    """Import one Excel sheet into a Postgres table via a live cursor.

    Returns the number of rows inserted.
    """
    excel_path = EXCEL_DIR / config["excel_file"]
    sheet_name = config["sheet"]
    table_name = config["table"]
    col_map = config["columns"]
    cols = list(col_map.keys())

    print(f"\n{'='*60}")
    print(f"  Table: {table_name}")
    print(f"  File:  {config['excel_file']} -> [{sheet_name}]")
    print(f"{'='*60}")

    if not excel_path.exists():
        print(f"  ERROR: Excel file not found: {excel_path}")
        return 0

    # Open Excel workbook in read-only mode
    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    except Exception as exc:
        print(f"  ERROR: Cannot open workbook: {exc}")
        return 0

    if sheet_name not in wb.sheetnames:
        print(f"  ERROR: Sheet '{sheet_name}' not found in {config['excel_file']}")
        print(f"  Available sheets: {wb.sheetnames}")
        wb.close()
        return 0

    ws = wb[sheet_name]

    # Read header row to build column index
    rows_iter = ws.iter_rows()
    try:
        header_row = next(rows_iter)
    except StopIteration:
        print("  WARNING: Sheet is empty")
        wb.close()
        return 0

    headers = [cell.value for cell in header_row]
    header_index = {}
    for i, h in enumerate(headers):
        if h is not None:
            header_index[str(h).strip()] = i

    # Verify all expected Excel columns exist
    missing_cols = []
    for db_col, excel_col in col_map.items():
        if excel_col not in header_index:
            missing_cols.append(f"{db_col}->{excel_col}")
    if missing_cols:
        print(f"  WARNING: Missing Excel columns: {', '.join(missing_cols)}")

    # Build FK lookup maps for any FK columns in this table
    fk_lookups = {}
    for db_col in col_map:
        key = (db_col, table_name)
        if key in FK_LOOKUPS:
            parent_file, parent_sheet, parent_name_col = FK_LOOKUPS[key]
            print(f"  Building FK lookup: {db_col} -> {parent_file}/{parent_sheet}")
            fk_lookups[db_col] = build_fk_lookup(parent_file, parent_sheet, parent_name_col)
            print(f"    -> {len(fk_lookups[db_col]):,} entries")

    # Read data rows
    batch = []
    total_rows = 0
    errors = 0

    for row in rows_iter:
        row_values = [cell.value for cell in row]

        # Skip completely empty rows
        if all(v is None for v in row_values):
            continue

        record = {}
        row_id = None

        for db_col, excel_col in col_map.items():
            col_idx = header_index.get(excel_col)
            if col_idx is None or col_idx >= len(row_values):
                record[db_col] = None
                continue

            raw_val = row_values[col_idx]

            # Handle FK columns: extract integer id (with name->id lookup)
            if db_col in FK_COLUMNS:
                record[db_col] = extract_fk_id(raw_val, fk_lookups.get(db_col))
            else:
                record[db_col] = flatten_value(raw_val)

            if db_col == "id":
                row_id = record[db_col]

        # Skip rows without an id
        if row_id is None:
            continue

        # Ensure id is an integer
        try:
            record["id"] = int(row_id)
        except (ValueError, TypeError):
            continue

        batch.append(record)
        total_rows += 1

        if len(batch) >= BATCH_SIZE:
            try:
                pg_insert(cur, table_name, cols, batch)
            except Exception as exc:
                print(f"  ERROR at row ~{total_rows}: {exc}")
                errors += 1
            batch = []

        # Progress reporting
        if total_rows % PROGRESS_EVERY == 0:
            print(f"  ... {total_rows:,} rows processed")

    if batch:
        try:
            pg_insert(cur, table_name, cols, batch)
        except Exception as exc:
            print(f"  ERROR in final batch: {exc}")
            errors += 1

    wb.close()

    status = "OK" if errors == 0 else f"with {errors} error(s)"
    print(f"  Done: {total_rows:,} rows -> {table_name} [{status}]")
    return total_rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    print("=" * 60)
    print("  Odoo Data Import to Supabase (fast / Postgres)")
    print("=" * 60)
    print(f"  Excel dir: {EXCEL_DIR}")
    print(f"  DB host:   {DATABASE_URL.split('@')[-1].split('/')[0] if '@' in DATABASE_URL else '(set)'}")
    print(f"  Batch size: {BATCH_SIZE}")
    print(f"  Tables to import: {len(IMPORT_MAP)}")
    print()

    # Validate directories exist
    if not EXCEL_DIR.exists():
        print(f"FATAL: Excel directory not found: {EXCEL_DIR}")
        sys.exit(1)

    # Connect to Postgres
    print("Connecting to Postgres ...", end=" ")
    try:
        conn = psycopg2.connect(DATABASE_URL, connect_timeout=30)
        conn.autocommit = False
        cur = conn.cursor()
        print("OK")
    except Exception as exc:
        print(f"FAILED: {exc}")
        sys.exit(1)

    # Wipe all tables — one instant TRUNCATE (the old REST delete-loop took ~45 min).
    # CASCADE covers the inter-table FKs; only these data tables reference each
    # other, so nothing outside the set is touched.
    all_tables = list(dict.fromkeys(c["table"] for c in IMPORT_MAP))
    print(f"\nTruncating {len(all_tables)} tables ...")
    try:
        cur.execute("TRUNCATE " + ", ".join(all_tables) + " CASCADE")
        conn.commit()
        print("Done wiping.\n")
    except Exception as exc:
        conn.rollback()
        print(f"FATAL: TRUNCATE failed: {exc}")
        sys.exit(1)

    # Import each sheet — commit per table so progress persists + FK lookups
    # (which re-read the parent sheet from Excel, not the DB) stay independent.
    start_time = time.time()
    summary = []

    for i, config in enumerate(IMPORT_MAP, 1):
        print(f"\n[{i}/{len(IMPORT_MAP)}]", end="")
        try:
            count = import_sheet(config, cur)
            conn.commit()
            summary.append((config["table"], count, "OK"))
        except Exception as exc:
            conn.rollback()
            print(f"  FATAL ERROR importing {config['table']}: {exc}")
            summary.append((config["table"], 0, f"ERROR: {exc}"))

    elapsed = time.time() - start_time
    cur.close()
    conn.close()

    # Print summary
    print("\n")
    print("=" * 60)
    print("  IMPORT SUMMARY")
    print("=" * 60)
    print(f"  {'Table':<30} {'Rows':>10}  Status")
    print(f"  {'-'*30} {'-'*10}  {'-'*20}")

    total_rows = 0
    ok_count = 0
    err_count = 0

    for table, count, status in summary:
        print(f"  {table:<30} {count:>10,}  {status}")
        total_rows += count
        if status == "OK":
            ok_count += 1
        else:
            err_count += 1

    print(f"  {'-'*30} {'-'*10}")
    print(f"  {'TOTAL':<30} {total_rows:>10,}")
    print()
    print(f"  Tables OK:     {ok_count}")
    print(f"  Tables ERROR:  {err_count}")
    print(f"  Time elapsed:  {elapsed:.1f}s")
    print("=" * 60)

    if err_count > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
